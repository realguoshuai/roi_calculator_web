#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ROI计算器 - 极速版
使用腾讯API获取实时价格，akshare获取财务和分红数据
"""

import sys
import os
import time
import requests
import logging
import json
from datetime import datetime
from roi import ROICalculator

# 设置编码
if sys.platform == 'win32':
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

# 获取脚本所在目录，用于导入同目录下的模块
if getattr(sys, 'frozen', False):
    script_dir = os.path.dirname(sys.executable)
else:
    script_dir = os.path.dirname(os.path.abspath(__file__))

# 添加脚本目录到sys.path
sys.path.insert(0, script_dir)

# 日志配置
log_dir = os.path.join(script_dir, "data", "log")
os.makedirs(log_dir, exist_ok=True)
log_file = os.path.join(log_dir, f"roi_fast_{datetime.now().strftime('%Y%m%d')}.log")

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(log_file, encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)


def get_stock_data_tencent(symbol: str) -> dict:
    """使用腾讯接口获取数据"""
    try:
        if symbol.startswith('SH'):
            code = 'sh' + symbol[2:]
        else:
            code = 'sz' + symbol[2:]
        
        url = f'https://qt.gtimg.cn/q={code}'
        response = requests.get(url, timeout=10, headers={
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
        })
        
        text = response.text.strip()
        parts = text.split('~')
        
        if len(parts) > 46:
            return {
                'name': parts[1],
                'price': float(parts[3]) if parts[3] else 0,
                'pe': float(parts[39]) if parts[39] else 0,
                'pb': float(parts[46]) if parts[46] else 0,
                'source': 'Tencent'
            }
        return None
    except Exception as e:
        return None


def get_custom_roe(symbol: str) -> float:
    """获取自定义ROE配置"""
    try:
        from config import CUSTOM_ROE
        return CUSTOM_ROE.get(symbol)
    except:
        return None


def get_financial_data_akshare(symbol: str) -> dict:
    """
    从akshare获取财务数据
    优先级：自定义ROE > 年度ROE(从年报数据中获取ROEJQ)
    """
    try:
        import akshare as ak
        
        if symbol.startswith('SH'):
            akshare_code = symbol[2:] + ".SH"
        else:
            akshare_code = symbol[2:] + ".SZ"
        
        df = ak.stock_financial_analysis_indicator_em(symbol=akshare_code)
        
        if len(df) > 0:
            # 1. 优先使用自定义ROE
            custom_roe = get_custom_roe(symbol)
            if custom_roe is not None:
                bps = float(df.iloc[0].get('BPS', 0)) if df.iloc[0].get('BPS') else 0
                return {
                    'roe': custom_roe,
                    'bps': bps,
                    'source': f'Custom({custom_roe}%)'
                }
            
            # 2. 从年报数据中获取年度ROE
            # REPORT_TYPE为"年报"的数据才是年度ROE
            annual_df = df[df['REPORT_TYPE'].str.contains('年报', na=False)]
            
            if len(annual_df) > 0:
                latest_annual = annual_df.iloc[0]
                bps = float(latest_annual.get('BPS', 0)) if latest_annual.get('BPS') else 0
                roe_annual = float(latest_annual.get('ROEJQ', 0)) if latest_annual.get('ROEJQ') else 0
                report_date = str(latest_annual.get('REPORT_DATE', '')) if latest_annual.get('REPORT_DATE') else ''
                
                if roe_annual > 0:
                    return {
                        'roe': roe_annual,
                        'bps': bps,
                        'source': f'Annual({report_date[:10]}, {roe_annual}%)'
                    }
            
            # 如果年报ROE为空，提示用户
            print(f"    [Warning] {symbol} 年度ROE数据为空")
            bps = float(df.iloc[0].get('BPS', 0)) if df.iloc[0].get('BPS') else 0
            return {
                'roe': 0,
                'bps': bps,
                'source': 'Annual(Empty)'
            }
        return None
    except Exception as e:
        print(f"    [Finance] Error: {e}")
        return None


def load_stocks() -> list:
    """
    加载股票列表
    优先级：
    1. 外部配置文件 stocks.json (exe同级目录)
    2. 内置 config.py
    """
    external_config = "stocks.json"
    stocks = None

    if os.path.exists(external_config):
        try:
            with open(external_config, 'r', encoding='utf-8') as f:
                stocks = json.load(f)
            print(f"[INFO] 读取外部配置文件: {external_config}")
            logger.info(f"Loaded external stocks from: {external_config}")
        except Exception as e:
            print(f"[WARN] 读取外部配置文件失败: {e}")
            logger.warning(f"Failed to load external config: {e}")

    if stocks is None:
        from config import STOCKS
        stocks = STOCKS
        print("[INFO] 使用内置股票配置")

    return stocks


def get_ttm_dividend_xq(symbol: str) -> dict:
    """
    从雪球接口获取TTM股息数据
    使用akshare的stock_individual_spot_xq接口
    """
    try:
        import akshare as ak
        
        # 雪球接口直接使用原始代码
        df = ak.stock_individual_spot_xq(symbol=symbol)
        
        # 转换为字典
        data = dict(zip(df['item'], df['value']))
        
        # 查找TTM股息和股息率
        ttm_dividend = 0
        ttm_yield = 0
        
        for item, value in data.items():
            if '股息(TTM)' in item:
                ttm_dividend = float(value) if value else 0
            elif '股息率(TTM)' in item:
                ttm_yield = float(value) if value else 0
        
        return {
            'ttm_dividend': round(ttm_dividend, 4),
            'ttm_yield': round(ttm_yield, 4),
            'source': 'Xueqiu(stock_individual_spot_xq)'
        }
    except Exception as e:
        print(f"    [TTM-Dividend] Error: {e}")
        return {
            'ttm_dividend': 0,
            'ttm_yield': 0,
            'source': f'Error: {e}'
        }


def load_stocks() -> list:
    """
    加载股票列表
    优先级：
    1. 外部配置文件 stocks.json (exe同级目录)
    2. 内置 config.py
    """
    external_config = "stocks.json"
    stocks = None

    if os.path.exists(external_config):
        try:
            with open(external_config, 'r', encoding='utf-8') as f:
                stocks = json.load(f)
            print(f"[INFO] 读取外部配置文件: {external_config}")
            logger.info(f"Loaded external stocks from: {external_config}")
        except Exception as e:
            print(f"[WARN] 读取外部配置文件失败: {e}")
            logger.warning(f"Failed to load external config: {e}")

    if stocks is None:
        from config import STOCKS
        stocks = STOCKS
        print("[INFO] 使用内置股票配置")

    return stocks


def get_custom_roe(symbol: str) -> float:
    """
    获取自定义ROE配置
    返回: 自定义ROE值(百分比)，如果未配置则返回None
    """
    try:
        from config import CUSTOM_ROE
        return CUSTOM_ROE.get(symbol)
    except:
        return None


def run_roi_analysis():
    """运行ROI分析程序"""
    STOCKS = load_stocks()
    
    logger.info("=" * 60)
    logger.info("ROI Calculator (Fast Version) Started")
    logger.info(f"Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    logger.info(f"Stocks: {len(STOCKS)}")
    logger.info("Data Source: Tencent (price) + akshare (finance + dividend)")
    logger.info("=" * 60)
    
    output_dir = "data/output"
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
    
    print("\n" + "=" * 60)
    print("  Investment ROI Calculator (Fast Version)")
    print(f"  Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 60)
    
    print(f"\nAnalyzing {len(STOCKS)} stocks")
    print("Data: Tencent (price) + akshare (finance + dividend)\n")
    
    start_time = time.time()
    calculator = ROICalculator()
    results = []
    
    for stock in STOCKS:
        name = stock['name']
        symbol = stock['symbol']
        print(f"Processing {name} ({symbol})...", end=" ", flush=True)
        
        try:
            # 价格
            data = get_stock_data_tencent(symbol)
            price = data['price'] if data else 0
            logger.info(f"[{name}] Price: {price}")
            
            # 财务数据
            fin_data = get_financial_data_akshare(symbol)
            roe = fin_data['roe'] if fin_data else 0
            # bps = fin_data['bps'] if fin_data else 0  # 保留BPS备用（已注释）
            # pb = round(price / bps, 2) if price > 0 and bps > 0 else 0  # 原计算方式（已注释）
            pb = data['pb'] if data else 0  # 直接使用腾讯API返回的PB
            logger.info(f"[{name}] ROE: {roe}%, PB: {pb}")
            
            # TTM股息数据(使用雪球接口)
            ttm_data = get_ttm_dividend_xq(symbol)
            ttm_dividend = ttm_data['ttm_dividend']
            ttm_yield = ttm_data['ttm_yield']
            logger.info(f"[{name}] TTM-Dividend(from Xueqiu): {ttm_dividend}, Yield: {ttm_yield}%")
            
            print(f"Price={price}, ROE={roe}%, PB={pb}")
            print(f"    TTM Dividend: {ttm_dividend}, Yield: {ttm_yield}%")
            
            stock_data = {
                'name': name,
                'symbol': symbol,
                'current_price': price,
                'financial': {'roe': roe, 'pb': pb},
                'dividend_yield': ttm_yield,
                'dividend': {'dividends': [{'cash_div': ttm_dividend}]}
            }
            
            result = calculator.calculate(stock_data)
            result.data_source = 'Tencent+Xueqiu'
            result.dividend_source = f'TTM({ttm_data["source"]})'
            result.pb_source = 'Tencent(qt.gtimg.cn)'
            results.append(result)
            
            logger.info(f"[{name}] Formula2-ROE/PB: {result.roi_formula2:.2f}%")
            
        except Exception as e:
            print(f"Failed: {e}")
            logger.error(f"[{name}] Error: {e}")
            continue
    
    elapsed = time.time() - start_time
    print(f"\nTime: {elapsed:.1f} seconds")
    logger.info(f"Execution Time: {elapsed:.1f} seconds")
    
    if results:
        print("\n" + "=" * 85)
        print("  ROI Comparison Summary")
        print("=" * 85)
        print(f"{'Name':<12} {'Code':<10} {'Price':<10} {'ROE':<8} {'PB':<8} {'LTM':<10} {'Yield':<12} {'ROE/PB':<10}")
        print("-" * 85)
        
        logger.info("=" * 85)
        logger.info("ROI Summary:")
        logger.info(f"{'Name':<12} {'Code':<10} {'Price':<10} {'ROE':<8} {'PB':<8} {'LTM':<10} {'F1-Yield':<12} {'F2-ROE/PB':<10}")
        logger.info("-" * 85)
        
        for r in results:
            price_str = f"{r.current_price:.2f}" if r.current_price else "N/A"
            roe_str = f"{r.roe:.2f}" if r.roe else "N/A"
            pb_str = f"{r.pb:.2f}" if r.pb else "N/A"
            ltm_str = f"{r.dividend_per_share:.4f}" if r.dividend_per_share else "N/A"
            yield_str = f"{r.roi_formula1:.2f}%" if r.roi_formula1 else "N/A"
            f2_str = f"{r.roi_formula2:.2f}%" if r.roi_formula2 else "N/A"
            print(f"{r.stock_name:<12} {r.symbol:<10} {price_str:<10} {roe_str:<8} {pb_str:<8} {ltm_str:<10} {yield_str:<12} {f2_str:<10}")
            logger.info(f"{r.stock_name:<12} {r.symbol:<10} {price_str:<10} {roe_str:<8} {pb_str:<8} {ltm_str:<10} {yield_str:<12} {f2_str:<10}")
        
        print("-" * 85)
        logger.info("-" * 85)
        
        # Yield ranking
        valid = [r for r in results if r.roi_formula1 and r.roi_formula1 > 0]
        sorted_yield = sorted(valid, key=lambda x: x.roi_formula1, reverse=True)
        print("\nFormula 1 (Dividend Yield) Ranking:")
        logger.info("\nFormula 1 (Dividend Yield) Ranking:")
        for i, r in enumerate(sorted_yield, 1):
            print(f"  {i}. {r.stock_name}: {r.roi_formula1:.2f}%")
            logger.info(f"  {i}. {r.stock_name}: {r.roi_formula1:.2f}%")
        
        # ROE/PB ranking
        valid_f2 = [r for r in results if r.roi_formula2 and r.roi_formula2 > 0]
        sorted_f2 = sorted(valid_f2, key=lambda x: x.roi_formula2, reverse=True)
        print("\nFormula 2 (ROE/PB) Ranking:")
        logger.info("\nFormula 2 (ROE/PB) Ranking:")
        for i, r in enumerate(sorted_f2, 1):
            print(f"  {i}. {r.stock_name}: {r.roi_formula2:.2f}%")
            logger.info(f"  {i}. {r.stock_name}: {r.roi_formula2:.2f}%")
        
        print("=" * 85)
        print("\nFormulas:")
        print("  F1 = (Dividend / Price) x 100%")
        print("  F2 = ROE / PB x 100%")
        print("  LTM = 2024 Annual - 2024 Mid + 2025 Mid")
        print("=" * 85)
        
        logger.info("=" * 85)
        logger.info("Formulas:")
        logger.info("  F1 = (Dividend / Price) x 100%")
        logger.info("  F2 = ROE / PB x 100%")
        logger.info("  LTM = 2024 Annual - 2024 Mid + 2025 Mid")
        logger.info("=" * 85)
        
        # 使用时间戳命名文件
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        save_to_excel(results, output_dir, timestamp)
        
        logger.info(f"Output files saved to: {output_dir}")
        logger.info("ROI Calculator (Fast Version) Completed")
        logger.info("=" * 60)


def save_to_excel(results, output_dir, timestamp=""):
    """保存结果到Excel（使用时间戳命名，直接保存在data目录）"""
    try:
        from openpyxl import Workbook
        
        # 使用时间戳命名文件，直接保存在data目录
        if timestamp:
            excel_path = os.path.join(output_dir, f"roi_{timestamp}.xlsx")
            chart_path = os.path.join(output_dir, f"roi_{timestamp}.png")
        else:
            excel_path = os.path.join(output_dir, "roi_analysis.xlsx")
            chart_path = os.path.join(output_dir, "roi_analysis.png")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "ROI Analysis"
        
        headers = ["Name", "Code", "Price", "ROE(%)", "PB", "LTM Dividend", "Yield(%)", "ROE/PB(%)", "Data Source"]
        
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        for row_idx, r in enumerate(results, 2):
            ws.cell(row=row_idx, column=1, value=r.stock_name)
            ws.cell(row=row_idx, column=2, value=r.symbol)
            ws.cell(row=row_idx, column=3, value=round(r.current_price, 2) if r.current_price else "N/A")
            ws.cell(row=row_idx, column=4, value=round(r.roe, 2) if r.roe else "N/A")
            ws.cell(row=row_idx, column=5, value=round(r.pb, 2) if r.pb else "N/A")
            ws.cell(row=row_idx, column=6, value=round(r.dividend_per_share, 4) if r.dividend_per_share else "N/A")
            ws.cell(row=row_idx, column=7, value=f"{r.roi_formula1:.2f}%" if r.roi_formula1 else "N/A")
            ws.cell(row=row_idx, column=8, value=f"{r.roi_formula2:.2f}%" if r.roi_formula2 else "N/A")
            ws.cell(row=row_idx, column=9, value=r.dividend_source)
        
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 8
        ws.column_dimensions['E'].width = 8
        ws.column_dimensions['F'].width = 14
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 12
        ws.column_dimensions['I'].width = 30
        
        wb.save(excel_path)
        print(f"\nExcel saved: {excel_path}")
        
        # 保存图表（生成3张PNG）
        save_chart(results, output_dir, timestamp)
        
    except Exception as e:
        print(f"\nExcel error: {e}")


def save_chart(results, output_dir, timestamp=""):
    """保存分析图表（生成3张PNG图片）
    - chart1: 口径1 (股息率) + ROE + Price + 分红详情
    - chart2: 口径2 (ROE/PB) + ROE + Price + PB
    - chart3: 口径1 + 口径2 汇总对比
    """
    try:
        import matplotlib
        matplotlib.use('Agg')
        import matplotlib.pyplot as plt
        
        plt.rcParams['font.sans-serif'] = ['SimHei', 'DejaVu Sans']
        plt.rcParams['axes.unicode_minus'] = False
        
        names = [r.stock_name for r in results]
        f1_values = [r.roi_formula1 if r.roi_formula1 else 0 for r in results]
        f2_values = [r.roi_formula2 if r.roi_formula2 else 0 for r in results]
        roes = [r.roe if r.roe else 0 for r in results]
        prices = [r.current_price if r.current_price else 0 for r in results]
        ltm_divs = [r.dividend_per_share if r.dividend_per_share else 0 for r in results]
        
        colors = ['#4472C4', '#ED7D31', '#70AD47', '#FFC000']
        
        # ========== Chart 1: 口径1 (股息率) Analysis ==========
        fig1, axes1 = plt.subplots(2, 2, figsize=(12, 10))
        fig1.suptitle(f'Analysis -口径1(股息率)- {datetime.now().strftime("%Y-%m-%d %H:%M")}', fontsize=14, fontweight='bold')
        
        # 口径1 (股息率)
        ax1 = axes1[0, 0]
        bars1 = ax1.bar(names, f1_values, color=colors[:len(names)])
        ax1.set_title('ROI-KouJing1: Dividend Yield (%)', fontsize=12, fontweight='bold')
        ax1.set_ylabel('Yield (%)')
        ax1.set_ylim(0, max(f1_values) * 1.3 if f1_values else 10)
        for bar, val in zip(bars1, f1_values):
            ax1.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 0.1, f'{val:.2f}%', ha='center', va='bottom')
        
        # ROE
        ax2 = axes1[0, 1]
        bars2 = ax2.bar(names, roes, color=colors[:len(names)])
        ax2.set_title('ROE (%)', fontsize=12, fontweight='bold')
        ax2.set_ylabel('ROE (%)')
        for bar, val in zip(bars2, roes):
            ax2.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 0.5, f'{val:.2f}%', ha='center', va='bottom')
        
        # Price
        ax3 = axes1[1, 0]
        bars3 = ax3.bar(names, prices, color=colors[:len(names)])
        ax3.set_title('Price (yuan)', fontsize=12, fontweight='bold')
        ax3.set_ylabel('Price (yuan)')
        for bar, val in zip(bars3, prices):
            ax3.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 5, f'{val:.2f}', ha='center', va='bottom')
        
        # LTM Dividend
        ax4 = axes1[1, 1]
        bars4 = ax4.bar(names, ltm_divs, color=colors[:len(names)])
        ax4.set_title('LTM Dividend (yuan)', fontsize=12, fontweight='bold')
        ax4.set_ylabel('Dividend (yuan)')
        for bar, val in zip(bars4, ltm_divs):
            ax4.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 0.1, f'{val:.4f}', ha='center', va='bottom')
        
        plt.tight_layout()
        chart1_path = os.path.join(output_dir, f"ROI_1_{timestamp}.png")
        fig1.savefig(chart1_path, dpi=150, bbox_inches='tight')
        print(f"Chart 1 saved: {chart1_path}")
        plt.close(fig1)
        
        # ========== Chart 2: 口径2 (ROE/PB) Analysis ==========
        fig2, axes2 = plt.subplots(2, 2, figsize=(12, 10))
        fig2.suptitle(f'Analysis -口径2(ROE_PB)- {datetime.now().strftime("%Y-%m-%d %H:%M")}', fontsize=14, fontweight='bold')
        
        # 口径2 (ROE/PB)
        ax1 = axes2[0, 0]
        bars1 = ax1.bar(names, f2_values, color=colors[:len(names)])
        ax1.set_title('ROI-KouJing2: ROE/PB (%)', fontsize=12, fontweight='bold')
        ax1.set_ylabel('ROE/PB (%)')
        ax1.set_ylim(0, max(f2_values) * 1.3 if f2_values else 10)
        for bar, val in zip(bars1, f2_values):
            ax1.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 0.1, f'{val:.2f}%', ha='center', va='bottom')
        
        # ROE
        ax2 = axes2[0, 1]
        bars2 = ax2.bar(names, roes, color=colors[:len(names)])
        ax2.set_title('ROE (%)', fontsize=12, fontweight='bold')
        ax2.set_ylabel('ROE (%)')
        for bar, val in zip(bars2, roes):
            ax2.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 0.5, f'{val:.2f}%', ha='center', va='bottom')
        
        # Price
        ax3 = axes2[1, 0]
        bars3 = ax3.bar(names, prices, color=colors[:len(names)])
        ax3.set_title('Price (yuan)', fontsize=12, fontweight='bold')
        ax3.set_ylabel('Price (yuan)')
        for bar, val in zip(bars3, prices):
            ax3.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 5, f'{val:.2f}', ha='center', va='bottom')
        
        # PB
        ax4 = axes2[1, 1]
        pbs = [r.pb if r.pb else 0 for r in results]
        bars4 = ax4.bar(names, pbs, color=colors[:len(names)])
        ax4.set_title('PB Ratio', fontsize=12, fontweight='bold')
        ax4.set_ylabel('PB')
        for bar, val in zip(bars4, pbs):
            ax4.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 0.1, f'{val:.2f}', ha='center', va='bottom')
        
        plt.tight_layout()
        chart2_path = os.path.join(output_dir, f"ROI_2_{timestamp}.png")
        fig2.savefig(chart2_path, dpi=150, bbox_inches='tight')
        print(f"Chart 2 saved: {chart2_path}")
        plt.close(fig2)
        
        # ========== Chart 3: 口径1 + 口径2 Combined ==========
        fig3, axes3 = plt.subplots(1, 2, figsize=(14, 6))
        fig3.suptitle(f'ROI Combined -口径1+口径2- {datetime.now().strftime("%Y-%m-%d %H:%M")}', fontsize=14, fontweight='bold')
        
        # 口径1 comparison
        ax1 = axes3[0]
        bars1 = ax1.bar(names, f1_values, color=colors[:len(names)], label='KouJing1')
        ax1.set_title('ROI-KouJing1: Dividend Yield (%)', fontsize=12, fontweight='bold')
        ax1.set_ylabel('Yield (%)')
        ax1.set_ylim(0, max(f1_values) * 1.3 if f1_values else 10)
        for bar, val in zip(bars1, f1_values):
            ax1.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 0.1, f'{val:.2f}%', ha='center', va='bottom')
        
        # 口径2 comparison
        ax2 = axes3[1]
        bars2 = ax2.bar(names, f2_values, color=colors[:len(names)], label='KouJing2')
        ax2.set_title('ROI-KouJing2: ROE/PB (%)', fontsize=12, fontweight='bold')
        ax2.set_ylabel('ROE/PB (%)')
        ax2.set_ylim(0, max(f2_values) * 1.3 if f2_values else 10)
        for bar, val in zip(bars2, f2_values):
            ax2.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 0.1, f'{val:.2f}%', ha='center', va='bottom')
        
        plt.tight_layout()
        chart3_path = os.path.join(output_dir, f"ROI_{timestamp}.png")
        fig3.savefig(chart3_path, dpi=150, bbox_inches='tight')
        print(f"Chart 3 saved: {chart3_path}")
        plt.close(fig3)
        
    except Exception as e:
        print(f"Chart error: {e}")


if __name__ == "__main__":
    run_roi_analysis()
