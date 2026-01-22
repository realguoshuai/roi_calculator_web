#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ROI计算器 - 增强版
功能：
1. 支持两种分红统计口径：最近12个月、年度分配
2. 动态获取分红和财务数据（无需配置）
3. 备注保底分红信息
数据来源：
- 价格: 腾讯实时API
- ROE/BPS: akshare stock_financial_analysis_indicator_em
- 分红: akshare stock_fhps_em (2024年度 + 2025中期)
"""

import sys
import os
import time
import requests
import logging
from datetime import datetime

# 设置编码
if sys.platform == 'win32':
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

project_dir = r"D:\code\git\roi_calculator"
sys.path.insert(0, project_dir)
os.chdir(project_dir)
sys.path.insert(0, project_dir)

# 日志配置
log_dir = os.path.join(project_dir, "data", "log")
os.makedirs(log_dir, exist_ok=True)
log_file = os.path.join(log_dir, f"roi_enhanced_{datetime.now().strftime('%Y%m%d')}.log")

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(log_file, encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)
os.chdir(project_dir)


def get_stock_data_tencent(symbol: str) -> dict:
    """使用腾讯接口获取实时价格数据"""
    try:
        if symbol.startswith('SH'):
            code = 'sh' + symbol[2:]
        else:
            code = 'sz' + symbol[2:]
        
        url = f'https://qt.gtimg.cn/q={code}'
        response = requests.get(url, timeout=10, headers={
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
        })
        
        text = response.text.strip()
        parts = text.split('~')
        
        if len(parts) > 45:
            return {
                'name': parts[1],
                'price': float(parts[3]) if parts[3] else 0,
                'pe': float(parts[37]) if parts[37] else 0,
                'pb': float(parts[38]) if parts[38] else 0,
                'source': '腾讯'
            }
        return None
    except Exception as e:
        return None


def get_custom_roe(symbol: str) -> float:
    """获取自定义ROE配置"""
    try:
        from config import CUSTOM_ROE
        return CUSTOM_ROE.get(symbol)
    except:
        return None


def get_financial_data_akshare(symbol: str) -> dict:
    """
    从akshare获取财务数据 (ROE, BPS)
    优先级：自定义ROE > 年度ROE(从年报数据中获取ROEJQ)
    """
    try:
        import akshare as ak
        
        if symbol.startswith('SH'):
            akshare_code = symbol[2:] + ".SH"
        else:
            akshare_code = symbol[2:] + ".SZ"
        
        df = ak.stock_financial_analysis_indicator_em(symbol=akshare_code)
        
        if len(df) > 0:
            # 1. 优先使用自定义ROE
            custom_roe = get_custom_roe(symbol)
            if custom_roe is not None:
                bps = float(df.iloc[0].get('BPS', 0)) if df.iloc[0].get('BPS') else 0
                report_date = str(df.iloc[0].get('REPORT_DATE', '')) if df.iloc[0].get('REPORT_DATE') else ''
                return {
                    'roe': custom_roe,
                    'bps': bps,
                    'report_date': report_date,
                    'source': f'Custom({custom_roe}%)'
                }
            
            # 2. 从年报数据中获取年度ROE
            # REPORT_TYPE为"年报"的数据才是年度ROE
            annual_df = df[df['REPORT_TYPE'].str.contains('年报', na=False)]
            
            if len(annual_df) > 0:
                latest_annual = annual_df.iloc[0]
                bps = float(latest_annual.get('BPS', 0)) if latest_annual.get('BPS') else 0
                roe_annual = float(latest_annual.get('ROEJQ', 0)) if latest_annual.get('ROEJQ') else 0
                report_date = str(latest_annual.get('REPORT_DATE', '')) if latest_annual.get('REPORT_DATE') else ''
                
                if roe_annual > 0:
                    return {
                        'roe': roe_annual,
                        'bps': bps,
                        'report_date': report_date,
                        'source': f'Annual({report_date[:10]}, {roe_annual}%)'
                    }
            
            # 如果年报ROE为空，提示用户
            print(f"    [Warning] {symbol} 年度ROE数据为空")
            bps = float(df.iloc[0].get('BPS', 0)) if df.iloc[0].get('BPS') else 0
            report_date = str(df.iloc[0].get('REPORT_DATE', '')) if df.iloc[0].get('REPORT_DATE') else ''
            return {
                'roe': 0,
                'bps': bps,
                'report_date': report_date,
                'source': 'Annual(Empty)'
            }
        return None
    except Exception as e:
        print(f"    [财务] 获取失败: {e}")
        return None


def get_dividend_data_akshare(symbol: str, price: float = 0) -> dict:
    """
    从akshare动态获取分红数据
    返回:
        interim_dividend: 2025年中期分红
        annual_dividend: 2024年度分红
        total_ltm: 最近12个月分红 (annual - interim_2024 + interim_2025)
        dividend_yield_ltm: LTM股息率(来自akshare)
        dividend_yield_annual: 年度股息率(来自akshare)
    说明:
        LTM = 2024年度 - 2024中期 + 2025中期
        因为2024年度分红已经包含了2024年中期分红
    """
    try:
        import akshare as ak
        
        interim_2025 = 0  # 2025年中期
        annual_2024 = 0   # 2024年度
        interim_2024 = 0  # 2024年中期（需要从年度中减去）
        interim_2025_yield = 0
        annual_2024_yield = 0
        
        # 获取2025年中期分红
        try:
            df = ak.stock_fhps_em(date='20250630')
            stock = df[df['代码'] == symbol[2:]]
            if len(stock) > 0:
                interim_2025 = float(stock.iloc[0].iloc[7]) if stock.iloc[0].iloc[7] else 0
                # Column 5 is already in percentage format (e.g., 12.70 = 12.70%)
                # Column 6 is decimal format (e.g., 0.0246 = 2.46%)
                # Use Column 5 directly as it's the dividend yield percentage
                interim_2025_yield = float(stock.iloc[0].iloc[5]) if stock.iloc[0].iloc[5] else 0
        except:
            pass
        
        # 获取2024年度分红
        try:
            df = ak.stock_fhps_em(date='20241231')
            stock = df[df['代码'] == symbol[2:]]
            if len(stock) > 0:
                annual_2024 = float(stock.iloc[0].iloc[7]) if stock.iloc[0].iloc[7] else 0
                # Column 5 is already in percentage format (e.g., 12.70 = 12.70%)
                annual_2024_yield = float(stock.iloc[0].iloc[5]) if stock.iloc[0].iloc[5] else 0
        except:
            pass
        
        # 获取2024年中期分红
        try:
            df = ak.stock_fhps_em(date='20240630')
            stock = df[df['代码'] == symbol[2:]]
            if len(stock) > 0:
                interim_2024 = float(stock.iloc[0].iloc[7]) if stock.iloc[0].iloc[7] else 0
        except:
            pass
        
        # LTM = 2024年度 - 2024中期 + 2025中期
        total_ltm = annual_2024 - interim_2024 + interim_2025
        
        # Calculate LTM dividend yield correctly: LTM dividend / current price × 100%
        # Akshare's column 5 is the dividend yield for that specific period, not LTM
        dividend_yield_ltm = (total_ltm / price * 100) if price > 0 and total_ltm > 0 else 0
        
        return {
            'interim_dividend': round(interim_2025, 4),
            'annual_dividend': round(annual_2024, 4),
            'interim_2024': round(interim_2024, 4),
            'total_ltm': round(total_ltm, 4),
            'dividend_yield_ltm': round(dividend_yield_ltm, 4),
            'dividend_yield_annual': round(annual_2024_yield, 4),
            'source': 'akshare(stock_fhps_em)'
        }
    except Exception as e:
        print(f"    [分红] 获取失败: {e}")
        return {
            'interim_dividend': 0,
            'annual_dividend': 0,
            'interim_2024': 0,
            'total_ltm': 0,
            'dividend_yield_ltm': 0,
            'dividend_yield_annual': 0,
            'source': f'Error: {e}'
        }


def get_guaranteed_dividend_note(symbol: str) -> str:
    """
    获取保底分红备注
    说明：保底分红信息通常在公司公告中披露，以下为常见公司的保底承诺
    需要定期从公司官网/公告中获取并更新
    """
    guaranteed_notes = {
        'SH600519': "【保底分红】贵州茅台：需查阅公司公告确认是否有未来三年保底分红承诺",
        'SZ000858': "【保底分红】五粮液：需查阅公司公告确认是否有未来三年保底分红承诺",
        'SZ000423': "【保底分红】东阿阿胶：需查阅公司公告确认是否有未来三年保底分红承诺",
        'SZ002304': "【保底分红】洋河股份：需查阅公司公告确认是否有未来三年保底分红承诺",
    }
    return guaranteed_notes.get(symbol, "【保底分红】需查阅公司公告确认")


def run_roi_analysis():
    """运行ROI分析程序"""
    from roi import ROICalculator
    
    # 股票列表
    stocks = [
        {"name": "东阿阿胶", "symbol": "SZ000423"},
        {"name": "五粮液", "symbol": "SZ000858"},
        {"name": "贵州茅台", "symbol": "SH600519"},
        {"name": "洋河股份", "symbol": "SZ002304"}
    ]
    
    print("\n" + "=" * 70)
    print("  投资回报率统计器 (增强版 - 动态数据)")
    print(f"  日期: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 70)
    
    logger.info("=" * 70)
    logger.info("ROI Calculator (Enhanced Version) Started")
    logger.info(f"Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    logger.info(f"Stocks: {len(stocks)}")
    logger.info("Data Source: Tencent (price) + akshare (finance + dividend)")
    logger.info("=" * 70)
    
    print(f"\n分析 {len(stocks)} 只股票")
    print("数据来源: 腾讯(价格) + akshare(财务+分红)")
    
    output_dir = "data/output"
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
    
    start_time = time.time()
    calculator = ROICalculator()
    results = []
    
    for stock in stocks:
        name = stock['name']
        symbol = stock['symbol']
        print(f"\n处理 {name} ({symbol})...", end=" ", flush=True)
        logger.info(f"[{name}] Processing...")
        
        try:
            # 1. 获取价格
            data = get_stock_data_tencent(symbol)
            price = data['price'] if data else 0
            logger.info(f"[{name}] Price: {price}")
            
            # 2. 获取财务数据 (ROE, BPS)
            fin_data = get_financial_data_akshare(symbol)
            roe = fin_data['roe'] if fin_data else 0
            bps = fin_data['bps'] if fin_data else 0
            pb = round(price / bps, 2) if price > 0 and bps > 0 else 0
            logger.info(f"[{name}] ROE: {roe}%, BPS: {bps}, PB: {pb}")
            
            # 3. 获取分红数据(传入股价，使用akshare提供的股息率)
            div_data = get_dividend_data_akshare(symbol, price)
            interim_2025 = div_data['interim_dividend']
            annual_2024 = div_data['annual_dividend']
            interim_2024 = div_data['interim_2024']
            total_ltm = div_data['total_ltm']
            dividend_yield_ltm = div_data['dividend_yield_ltm']
            dividend_yield_annual = div_data['dividend_yield_annual']
            logger.info(f"[{name}] Dividends - Annual2024: {annual_2024}, Mid2024: {interim_2024}, Mid2025: {interim_2025}")
            logger.info(f"[{name}] Dividend-Yield(from akshare): LTM={dividend_yield_ltm:.2f}%, Annual={dividend_yield_annual:.2f}%")
            
            print(f"价格={price}, ROE={roe}%, PB={pb}")
            print(f"    分红: 年度={annual_2024}, 2024中期={interim_2024}, 2025中期={interim_2025}")
            print(f"    LTM={total_ltm} ({annual_2024}-{interim_2024}+{interim_2025})")
            print(f"    股息率: LTM={dividend_yield_ltm:.2f}%, 年度={dividend_yield_annual:.2f}%")
            
            # 创建两个结果对象
            # 结果1: 使用LTM分红
            stock_data_ltm = {
                'name': name,
                'symbol': symbol,
                'current_price': price,
                'financial': {'roe': roe, 'pb': pb},
                'dividend_yield': dividend_yield_ltm,
                'dividend': {'dividends': [{'cash_div': total_ltm}]}
            }
            
            result_ltm = calculator.calculate(stock_data_ltm)
            result_ltm.data_source = '腾讯+akshare'
            result_ltm.dividend_source = f"LTM({annual_2024}-{interim_2024}+{interim_2025})"
            result_ltm.pb_source = f'akshare(Price={price}/BPS={bps:.2f})'
            result_ltm.guaranteed_note = get_guaranteed_dividend_note(symbol)
            result_ltm.interim_dividend = interim_2025
            result_ltm.annual_dividend = annual_2024
            results.append(result_ltm)
            
            # 结果2: 使用年度分红
            stock_data_annual = {
                'name': name,
                'symbol': symbol,
                'current_price': price,
                'financial': {'roe': roe, 'pb': pb},
                'dividend_yield': dividend_yield_annual,
                'dividend': {'dividends': [{'cash_div': annual_2024}]}
            }
            
            result_annual = calculator.calculate(stock_data_annual)
            result_annual.data_source = '腾讯+akshare'
            result_annual.dividend_source = f"年度分配({annual_2024})"
            result_annual.pb_source = f'akshare(Price={price}/BPS={bps:.2f})'
            result_annual.guaranteed_note = get_guaranteed_dividend_note(symbol)
            result_annual.interim_dividend = 0
            result_annual.annual_dividend = annual_2024
            results.append(result_annual)
            
            logger.info(f"[{name}] KouJing1(LTM): F1={result_ltm.roi_formula1:.2f}%, F2={result_ltm.roi_formula2:.2f}%")
            logger.info(f"[{name}] KouJing2(Annual): F1={result_annual.roi_formula1:.2f}%, F2={result_annual.roi_formula2:.2f}%")
            
        except Exception as e:
            print(f"处理失败: {e}")
            logger.error(f"[{name}] Error: {e}")
            continue
        
        elapsed = time.time() - start_time
        print(f"\n数据获取耗时: {elapsed:.1f}秒")
        logger.info(f"Data fetch time: {elapsed:.1f} seconds")
        
        if results:
            logger.info("=" * 100)
            logger.info("ROI Summary")
            logger.info("=" * 100)
            
            print("\n" + "=" * 100)
            print("  投资回报率对比汇总")
            print("=" * 100)
            
            # 按类型分组显示
            ltm_results = [r for r in results if 'LTM' in r.dividend_source]
            annual_results = [r for r in results if '年度分配' in r.dividend_source]
            
            # 显示LTM结果
            print("\n【口径一：最近12个月分红 (2024年度 + 2025中期)】")
            logger.info("\n[KouJing1: LTM]")
            print(f"{'股票名称':<12} {'代码':<10} {'股价':<10} {'ROE':<8} {'PB':<8} {'分红':<10} {'股息率':<10} {'ROE/PB':<10}")
            print("-" * 90)
            logger.info(f"{'Name':<12} {'Code':<10} {'Price':<10} {'ROE':<8} {'PB':<8} {'Div':<10} {'F1-Yield':<10} {'F2-ROE/PB':<10}")
            logger.info("-" * 90)
            
            valid_ltm = [r for r in ltm_results if r.roi_formula1 and r.roi_formula1 > 0]
            sorted_ltm = sorted(valid_ltm, key=lambda x: x.roi_formula1, reverse=True)
            for r in sorted_ltm:
                print(f"{r.stock_name:<12} {r.symbol:<10} {r.current_price:<10.2f} {r.roe:<8.2f} {r.pb:<8.2f} {r.dividend_per_share:<10.4f} {r.roi_formula1:<10.2f}% {r.roi_formula2:<10.2f}%")
                logger.info(f"{r.stock_name:<12} {r.symbol:<10} {r.current_price:<10.2f} {r.roe:<8.2f} {r.pb:<8.2f} {r.dividend_per_share:<10.4f} {r.roi_formula1:<10.2f}% {r.roi_formula2:<10.2f}%")
            
            if sorted_ltm:
                print("\n  股息率排名(LTM):")
                logger.info("\n  Ranking (F1 Yield):")
                for i, r in enumerate(sorted_ltm, 1):
                    print(f"    {i}. {r.stock_name}: {r.roi_formula1:.2f}%")
                    logger.info(f"    {i}. {r.stock_name}: {r.roi_formula1:.2f}%")
        
            # 显示年度分配结果
            print("\n【口径二：年度分配分红 (仅2024年度)】")
            logger.info("\n[KouJing2: Annual]")
            print(f"{'股票名称':<12} {'代码':<10} {'股价':<10} {'ROE':<8} {'PB':<8} {'分红':<10} {'股息率':<10} {'ROE/PB':<10}")
            print("-" * 90)
            logger.info(f"{'Name':<12} {'Code':<10} {'Price':<10} {'ROE':<8} {'PB':<8} {'Div':<10} {'F1-Yield':<10} {'F2-ROE/PB':<10}")
            logger.info("-" * 90)
        
            valid_annual = [r for r in annual_results if r.roi_formula1 and r.roi_formula1 > 0]
            sorted_annual = sorted(valid_annual, key=lambda x: x.roi_formula1, reverse=True)
            for r in sorted_annual:
                print(f"{r.stock_name:<12} {r.symbol:<10} {r.current_price:<10.2f} {r.roe:<8.2f} {r.pb:<8.2f} {r.dividend_per_share:<10.4f} {r.roi_formula1:<10.2f}% {r.roi_formula2:<10.2f}%")
                logger.info(f"{r.stock_name:<12} {r.symbol:<10} {r.current_price:<10.2f} {r.roe:<8.2f} {r.pb:<8.2f} {r.dividend_per_share:<10.4f} {r.roi_formula1:<10.2f}% {r.roi_formula2:<10.2f}%")
        
            if sorted_annual:
                print("\n  股息率排名(年度):")
                logger.info("\n  Ranking (F1 Yield):")
                for i, r in enumerate(sorted_annual, 1):
                    print(f"    {i}. {r.stock_name}: {r.roi_formula1:.2f}%")
                    logger.info(f"    {i}. {r.stock_name}: {r.roi_formula1:.2f}%")
        
            # ROE/PB排名
            print("\n【公式二：ROE/PB 排名】")
            logger.info("\n[Formula 2: ROE/PB Ranking]")
            all_results = ltm_results  # 使用LTM结果
            valid_f2 = [r for r in all_results if r.roi_formula2 and r.roi_formula2 > 0]
            sorted_f2 = sorted(valid_f2, key=lambda x: x.roi_formula2, reverse=True)
            for i, r in enumerate(sorted_f2, 1):
                print(f"    {i}. {r.stock_name}: {r.roi_formula2:.2f}%")
                logger.info(f"    {i}. {r.stock_name}: {r.roi_formula2:.2f}%")
        
            print("\n" + "=" * 100)
        
            # 保存结果（使用时间戳命名）
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            save_to_excel(ltm_results, annual_results, output_dir, timestamp)
        
            # 打印备注
            print("\n【保底分红备注】")
            logger.info("\n[Guaranteed Dividend Notes]")
            for stock in stocks:
                symbol = stock['symbol']
                ltm_r = next((r for r in ltm_results if r.symbol == symbol), None)
                if ltm_r:
                    print(f"  {stock['name']}: {ltm_r.guaranteed_note}")
                    logger.info(f"  {stock['name']}: {ltm_r.guaranteed_note}")
        
            print("\n" + "=" * 100)
            print("计算公式说明:")
            print("  公式一 = (分红 / 当前股价) × 100%")
            print("  公式二 = ROE / PB × 100% (注: ROE不带百分号直接相除)")
            print("  LTM = 最近12个月分红 = 2024年度分红 + 2025年中期分红")
            print("=" * 100)
        
            logger.info("=" * 100)
            logger.info("Formulas:")
            logger.info("  F1 = (Dividend / Price) x 100%")
            logger.info("  F2 = ROE / PB x 100%")
            logger.info("  LTM = 2024 Annual + 2025 Mid")
            logger.info("=" * 100)
            logger.info(f"Output files saved to: {output_dir}")
            logger.info("ROI Calculator (Enhanced Version) Completed")
            logger.info("=" * 70)


def save_to_excel(ltm_results, annual_results, output_dir, timestamp=""):
    """保存结果到Excel（使用时间戳命名，直接保存在data目录）"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        
        # 使用时间戳命名文件，直接保存在data目录
        if timestamp:
            excel_path = os.path.join(output_dir, f"roi_{timestamp}.xlsx")
            chart_path = os.path.join(output_dir, f"roi_{timestamp}.png")
        else:
            excel_path = os.path.join(output_dir, "roi_analysis.xlsx")
            chart_path = os.path.join(output_dir, "roi_analysis.png")
        
        wb = Workbook()
        
        # LTM结果表
        ws1 = wb.active
        ws1.title = "LTM分红"
        
        headers1 = ["股票名称", "代码", "股价(元)", "ROE(%)", "PB", "年度分红", "中期分红", "LTM分红", "股息率(%)", "ROE/PB(%)", "保底分红备注"]
        for col, header in enumerate(headers1, 1):
            cell = ws1.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        for row_idx, r in enumerate(ltm_results, 2):
            ws1.cell(row=row_idx, column=1, value=r.stock_name)
            ws1.cell(row=row_idx, column=2, value=r.symbol)
            ws1.cell(row=row_idx, column=3, value=round(r.current_price, 2))
            ws1.cell(row=row_idx, column=4, value=round(r.roe, 2))
            ws1.cell(row=row_idx, column=5, value=round(r.pb, 2))
            ws1.cell(row=row_idx, column=6, value=r.annual_dividend)
            ws1.cell(row=row_idx, column=7, value=r.interim_dividend)
            ws1.cell(row=row_idx, column=8, value=r.dividend_per_share)
            ws1.cell(row=row_idx, column=9, value=f"{r.roi_formula1:.2f}%")
            ws1.cell(row=row_idx, column=10, value=f"{r.roi_formula2:.2f}%")
            ws1.cell(row=row_idx, column=11, value=r.guaranteed_note)
        
        # 年度分配结果表
        ws2 = wb.create_sheet("年度分配")
        
        headers2 = ["股票名称", "代码", "股价(元)", "ROE(%)", "PB", "年度分红(元)", "股息率(%)", "ROE/PB(%)", "保底分红备注"]
        for col, header in enumerate(headers2, 1):
            cell = ws2.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        for row_idx, r in enumerate(annual_results, 2):
            ws2.cell(row=row_idx, column=1, value=r.stock_name)
            ws2.cell(row=row_idx, column=2, value=r.symbol)
            ws2.cell(row=row_idx, column=3, value=round(r.current_price, 2))
            ws2.cell(row=row_idx, column=4, value=round(r.roe, 2))
            ws2.cell(row=row_idx, column=5, value=round(r.pb, 2))
            ws2.cell(row=row_idx, column=6, value=r.annual_dividend)
            ws2.cell(row=row_idx, column=7, value=f"{r.roi_formula1:.2f}%")
            ws2.cell(row=row_idx, column=8, value=f"{r.roi_formula2:.2f}%")
            ws2.cell(row=row_idx, column=9, value=r.guaranteed_note)
        
        # 调整列宽
        for ws in [ws1, ws2]:
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 10
            ws.column_dimensions['C'].width = 10
            ws.column_dimensions['D'].width = 8
            ws.column_dimensions['E'].width = 8
            ws.column_dimensions['F'].width = 12
            ws.column_dimensions['G'].width = 12
            ws.column_dimensions['H'].width = 12
            ws.column_dimensions['I'].width = 12
            ws.column_dimensions['J'].width = 12
            ws.column_dimensions['K'].width = 50 if ws == ws1 else 50
        
        wb.save(excel_path)
        print(f"\nExcel saved: {excel_path}")
        
        # 保存图表
        save_enhanced_chart(ltm_results, annual_results, chart_path)
        
    except Exception as e:
        print(f"\n保存Excel失败: {e}")


def save_enhanced_chart(ltm_results, annual_results, output_dir, timestamp=""):
    """保存分析图表（生成3张PNG图片）
    - chart1: 口径1 (股息率) + ROE + Price + 分红
    - chart2: 口径2 (ROE/PB) + ROE + Price + PB
    - chart3: 口径1 + 口径2 汇总对比
    """
    try:
        import matplotlib
        matplotlib.use('Agg')
        import matplotlib.pyplot as plt
        
        plt.rcParams['font.sans-serif'] = ['SimHei', 'DejaVu Sans']
        plt.rcParams['axes.unicode_minus'] = False
        
        ltm_names = [r.stock_name for r in ltm_results]
        ltm_f1 = [r.roi_formula1 if r.roi_formula1 else 0 for r in ltm_results]
        ltm_f2 = [r.roi_formula2 if r.roi_formula2 else 0 for r in ltm_results]
        roes = [r.roe if r.roe else 0 for r in ltm_results]
        prices = [r.current_price if r.current_price else 0 for r in ltm_results]
        ltm_divs = [r.dividend_per_share if r.dividend_per_share else 0 for r in ltm_results]
        pbs = [r.pb if r.pb else 0 for r in ltm_results]
        
        colors = ['#4472C4', '#ED7D31', '#70AD47', '#FFC000']
        
        # ========== Chart 1: 口径1 (股息率) Analysis ==========
        fig1, axes1 = plt.subplots(2, 2, figsize=(12, 10))
        fig1.suptitle(f'Analysis -口径1(股息率)- {datetime.now().strftime("%Y-%m-%d %H:%M")}', fontsize=14, fontweight='bold')
        
        # 口径1
        ax1 = axes1[0, 0]
        bars1 = ax1.bar(ltm_names, ltm_f1, color=colors[:len(ltm_names)])
        ax1.set_title('ROI-KouJing1: Dividend Yield (%)', fontsize=12, fontweight='bold')
        ax1.set_ylabel('Yield (%)')
        ax1.set_ylim(0, max(ltm_f1) * 1.3 if ltm_f1 else 10)
        for bar, val in zip(bars1, ltm_f1):
            ax1.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 0.1, f'{val:.2f}%', ha='center', va='bottom')
        
        # ROE
        ax2 = axes1[0, 1]
        bars2 = ax2.bar(ltm_names, roes, color=colors[:len(ltm_names)])
        ax2.set_title('ROE (%)', fontsize=12, fontweight='bold')
        ax2.set_ylabel('ROE (%)')
        for bar, val in zip(bars2, roes):
            ax2.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 0.5, f'{val:.2f}%', ha='center', va='bottom')
        
        # Price
        ax3 = axes1[1, 0]
        bars3 = ax3.bar(ltm_names, prices, color=colors[:len(ltm_names)])
        ax3.set_title('Price (yuan)', fontsize=12, fontweight='bold')
        ax3.set_ylabel('Price (yuan)')
        for bar, val in zip(bars3, prices):
            ax3.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 5, f'{val:.2f}', ha='center', va='bottom')
        
        # LTM Dividend
        ax4 = axes1[1, 1]
        bars4 = ax4.bar(ltm_names, ltm_divs, color=colors[:len(ltm_names)])
        ax4.set_title('LTM Dividend (yuan)', fontsize=12, fontweight='bold')
        ax4.set_ylabel('Dividend (yuan)')
        for bar, val in zip(bars4, ltm_divs):
            ax4.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 0.1, f'{val:.4f}', ha='center', va='bottom')
        
        plt.tight_layout()
        chart1_path = os.path.join(output_dir, f"ROI_1_{timestamp}.png")
        fig1.savefig(chart1_path, dpi=150, bbox_inches='tight')
        print(f"Chart 1 saved: {chart1_path}")
        plt.close(fig1)
        
        # ========== Chart 2: 口径2 (ROE/PB) Analysis ==========
        fig2, axes2 = plt.subplots(2, 2, figsize=(12, 10))
        fig2.suptitle(f'Analysis -口径2(ROE_PB)- {datetime.now().strftime("%Y-%m-%d %H:%M")}', fontsize=14, fontweight='bold')
        
        # 口径2
        ax1 = axes2[0, 0]
        bars1 = ax1.bar(ltm_names, ltm_f2, color=colors[:len(ltm_names)])
        ax1.set_title('ROI-KouJing2: ROE/PB (%)', fontsize=12, fontweight='bold')
        ax1.set_ylabel('ROE/PB (%)')
        ax1.set_ylim(0, max(ltm_f2) * 1.3 if ltm_f2 else 10)
        for bar, val in zip(bars1, ltm_f2):
            ax1.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 0.1, f'{val:.2f}%', ha='center', va='bottom')
        
        # ROE
        ax2 = axes2[0, 1]
        bars2 = ax2.bar(ltm_names, roes, color=colors[:len(ltm_names)])
        ax2.set_title('ROE (%)', fontsize=12, fontweight='bold')
        ax2.set_ylabel('ROE (%)')
        for bar, val in zip(bars2, roes):
            ax2.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 0.5, f'{val:.2f}%', ha='center', va='bottom')
        
        # Price
        ax3 = axes2[1, 0]
        bars3 = ax3.bar(ltm_names, prices, color=colors[:len(ltm_names)])
        ax3.set_title('Price (yuan)', fontsize=12, fontweight='bold')
        ax3.set_ylabel('Price (yuan)')
        for bar, val in zip(bars3, prices):
            ax3.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 5, f'{val:.2f}', ha='center', va='bottom')
        
        # PB
        ax4 = axes2[1, 1]
        bars4 = ax4.bar(ltm_names, pbs, color=colors[:len(ltm_names)])
        ax4.set_title('PB Ratio', fontsize=12, fontweight='bold')
        ax4.set_ylabel('PB')
        for bar, val in zip(bars4, pbs):
            ax4.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 0.1, f'{val:.2f}', ha='center', va='bottom')
        
        plt.tight_layout()
        chart2_path = os.path.join(output_dir, f"ROI_2_{timestamp}.png")
        fig2.savefig(chart2_path, dpi=150, bbox_inches='tight')
        print(f"Chart 2 saved: {chart2_path}")
        plt.close(fig2)
        
        # ========== Chart 3: 口径1 + 口径2 Combined ==========
        fig3, axes3 = plt.subplots(1, 2, figsize=(14, 6))
        fig3.suptitle(f'ROI Combined -口径1+口径2- {datetime.now().strftime("%Y-%m-%d %H:%M")}', fontsize=14, fontweight='bold')
        
        # 口径1
        ax1 = axes3[0]
        bars1 = ax1.bar(ltm_names, ltm_f1, color=colors[:len(ltm_names)], label='KouJing1')
        ax1.set_title('ROI-KouJing1: Dividend Yield (%)', fontsize=12, fontweight='bold')
        ax1.set_ylabel('Yield (%)')
        ax1.set_ylim(0, max(ltm_f1) * 1.3 if ltm_f1 else 10)
        for bar, val in zip(bars1, ltm_f1):
            ax1.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 0.1, f'{val:.2f}%', ha='center', va='bottom')
        
        # 口径2
        ax2 = axes3[1]
        bars2 = ax2.bar(ltm_names, ltm_f2, color=colors[:len(ltm_names)], label='KouJing2')
        ax2.set_title('ROI-KouJing2: ROE/PB (%)', fontsize=12, fontweight='bold')
        ax2.set_ylabel('ROE/PB (%)')
        ax2.set_ylim(0, max(ltm_f2) * 1.3 if ltm_f2 else 10)
        for bar, val in zip(bars2, ltm_f2):
            ax2.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 0.1, f'{val:.2f}%', ha='center', va='bottom')
        
        plt.tight_layout()
        chart3_path = os.path.join(output_dir, f"ROI_{timestamp}.png")
        fig3.savefig(chart3_path, dpi=150, bbox_inches='tight')
        print(f"Chart 3 saved: {chart3_path}")
        plt.close(fig3)
        
    except Exception as e:
        print(f"Chart error: {e}")


if __name__ == "__main__":
    run_roi_analysis()
